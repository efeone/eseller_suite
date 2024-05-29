import frappe
from io import BytesIO
import openpyxl
import xlrd
from openpyxl import load_workbook

def read_xlsx_file_from_attached_file(file_url=None, fcontent=None, filepath=None):
    if file_url:
        _file = frappe.get_doc("File", {"file_url": file_url})
        filename = _file.get_full_path()
    elif fcontent:
        filename = BytesIO(fcontent)
    elif filepath:
        filename = filepath
    else:
        return

    data = {}
    wb1 = load_workbook(filename=filename, data_only=True)
    print("\n\n\n\n\n\n\n\n\n\n\n")
    print(wb1)
    print(wb1.sheetnames)
    for sheet_name in wb1.sheetnames:
        rows = []
        ws1 = wb1[sheet_name]
        for row in ws1.iter_rows():
            rows.append([cell.value for cell in row])
        data[sheet_name] = rows
    print("\n\n\n\n\n\n\n\n\n\n\n")
    # print(data)
    return data

def read_xls_file_from_attached_file(content):
	book = xlrd.open_workbook(file_contents=content)
	sheets = book.sheets()
	sheet = sheets[0]
	return [sheet.row_values(i) for i in range(sheet.nrows)]
